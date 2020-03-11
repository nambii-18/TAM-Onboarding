import os
import shutil
from openpyxl import load_workbook
import sys
import mysql.connector as mysql
# import time
user = 'root'

tam_dir = '/app/TAM/'
# template_dir = '/app/tam-template/'
# files = os.listdir(template_dir)
# os.mkdir(dir)

os.system('cd /app && git clone -b master https://github.com/nambii-18/TAM-Onboarding.git')
# os.rename(dir+'/TAM-Onboarding',dir+'/'+sys.argv[1])

shutil.copytree('/app/TAM-Onboarding/tam-template',tam_dir+'/'+sys.argv[1])

tam_dir = tam_dir + '/'  +sys.argv[1] + '/'
# # print(files)
# for file in files:
#     try:
#         shutil.copy(template_dir+file,dir+file)
#     except Exception as e:
#         print(e)
#         if os.path.isdir(template_dir+file):
#             shutil.copytree(template_dir+file,dir+file)

with open(tam_dir+'connect-hsbc.php') as f:
  lines = f.readlines()
  lines[2] = lines[2].split('_')[0] + '_' + sys.argv[1] + "');\n"
  lines = ''.join(lines)
  # print(lines)

with open(tam_dir+'connect-hsbc.php','w') as f:
  f.write(lines)


db = mysql.connect(
    host = "localhost",
    user = user
)

db_name = 'tam_'+sys.argv[1]
cursor = db.cursor()
cursor.execute('CREATE DATABASE '+db_name)
cursor.execute('USE '+db_name+';')


with open('/app/mysql-template/scripts/tam-hsbc.sql') as f:
  lines = f.read().split(';')

# print(lines)
for line in lines:
  cursor.execute(line)

wb = load_workbook('/app/upload/OnboardExcel.xlsx')
sheets = ['Engineers','Reviewers','Customers','Managers or External Access']

sheet = wb['Engineers']
rows = sheet.rows
for i,row in enumerate(rows):
  if i!= 0:
    cec = row[0].value
    gdc = row[1].value
    cursor.execute('INSERT INTO c_eng VALUES ' + "('"+cec+"','"+gdc+"',0,0,0,0,0,0,0,0,0,0)")

sheet = wb['Customers']
rows = sheet.rows
for i,row in enumerate(rows):
  if i!= 0:
    customer = row[0].value
    cursor.execute('INSERT INTO customer VALUES ' + "('"+customer+"','NULL')")

if(len(sys.argv) == 3):
  crt_dir = '/app/CRT/'+sys.argv[1]
  shutil.copytree('/app/TAM-Onboarding/crt-template',crt_dir)

  with open(crt_dir+'/connect.php') as f:
    lines = f.readlines()
    lines[2] = lines[2].split('_')[0] + '_' + sys.argv[1] + "');\n"
    lines = ''.join(lines)
    # print(lines)

  with open(crt_dir+'/connect.php','w') as f:
    f.write(lines)


  # db = mysql.connect(
  #     host = "localhost",
  #     user = user
  # )

  db_name = 'crt_'+sys.argv[1]
  # cursor = db.cursor()
  cursor.execute('CREATE DATABASE '+db_name)
  cursor.execute('USE '+db_name+';')


  with open('/app/mysql-template/scripts/crt-hsbc.sql') as f:
    lines = f.read().split(';')

  print(lines)
  
  for line in lines:
    try:
      cursor.execute(line)
    except:
      pass

  wb = load_workbook('/app/upload/OnboardExcel.xlsx')
  sheets = ['Engineers','Reviewers','Customers','Managers or External Access']

  sheet = wb['Engineers']
  rows = sheet.rows
  for i,row in enumerate(rows):
    if i!= 0:
      cec = row[0].value
      gdc = row[1].value
      manager = row[2].value
      cursor.execute('INSERT INTO engineer VALUES ' + "('"+cec+"','"+manager+"','"+gdc+"')")

  sheet = wb['Customers']
  rows = sheet.rows
  for i,row in enumerate(rows):
    if i!= 0:
      customer = row[0].value
      cursor.execute('INSERT INTO customer VALUES ' + "('"+customer+"')")

  sheet = wb['Managers or External Access']
  rows = sheet.rows
  for i,row in enumerate(rows):
    if i!=0:
      manager = row[0].value
      gdc = row[1].value
      cursor.execute('INSERT INTO manager VALUES ' + "('"+manager+"','"+gdc+"')")

  sheet = wb['Reviewers']
  rows = sheet.rows
  for i,row in enumerate(rows):
    if i!=0:
      reviewer = row[0].value
      gdc = row[1].value
      cursor.execute('INSERT INTO reviewers (CEC,GDC) VALUES ' + "('"+reviewer+"','"+gdc+"')")
      cursor.execute('INSERT INTO reviewers2 (CEC,GDC) VALUES ' + "('"+reviewer+"','"+gdc+"')")

shutil.rmtree('/app/TAM-Onboarding')
cursor.close()
db.close()