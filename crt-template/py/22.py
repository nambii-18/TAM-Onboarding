#import win32com
from openpyxl import Workbook
from openpyxl import load_workbook
import re
from openpyxl.styles import NamedStyle, Font, colors, Border, Side
from openpyxl.styles.numbers import NumberFormat
#import win32com.client as win32
wb = Workbook()
import sys
import pandas as pd



filename = sys.argv[1]
#print filename + "sdssd"
#load sheets
wb2 = load_workbook('/var/www/html/CRT/HSBC/py/'+filename)
names = wb2.get_sheet_names()
ws = wb2.get_sheet_by_name(names[0])
print ws.max_row

wb_map = load_workbook('/var/www/html/CRT/HSBC/py/GSD to CEC Mapping.xlsx')
print wb_map.get_sheet_names()
ws_map = wb_map.get_sheet_by_name('names')

#set style
date_style = NamedStyle(name='datetime', number_format='MM/DD/YYYY')
#here you iterate over the rows in the specific column
#print([re.sub(u'(\u2018|\u2019)', "'", str(ws[x][5].value)) for x in range(2,ws.max_row)])
'''
B = ws['B']
print B[1].value
ft = Font(color=colors.RED)
B[1].font = ft
B[1].value = "Changed by me"
'''

#format columns - important
for x in range(2,ws.max_row+1):
	#print x,ws['C' + str(x)].value
	try:
		ws['C' + str(x)].value=ws['C' + str(x)].value+", "+ws['B'+str(x)].value
		ws['B' + str(x)].value = ws['A' + str(x)].value
		ws['A' + str(x)].value="HSBC"
		ws['D' + str(x)].value=ws['J'+str(x)].value

		ws['F' + str(x)].value=ws['C'+str(x)].value
		ws['C' + str(x)].value=ws['I'+str(x)].value
		ws['E' + str(x)].style = date_style
		#change the value itself
		#ws['E' + str(x)].style = date_style

		ws['G' + str(x)].value=" "
		ws['H' + str(x)].value=" "
		ws['I' + str(x)].value=" "
		ws['J' + str(x)].value=" "
		ws['K' + str(x)].value=" "
		ws['L' + str(x)].value=" "
		ws['M' + str(x)].value=" "
		ws['N' + str(x)].value=" "
		ws['O' + str(x)].value=" "
	except Exception, e:
		print e
	finally:
		pass

# hsbc in sta pri date cec
#modify F now into CEC's
for i in range(1,ws_map.max_row+1):
	for x in range(2,ws.max_row+1):
		try:
			#print "map=",ws_map['A' + str(i)].value
			if ws['F' + str(x)].value.replace(" ", "") == ws_map['A' + str(i)].value.replace(" ", ""):
				ws['F' + str(x)].value = ws_map['B' + str(i)].value
		except Exception,e:
			print e
		finally:
			pass

#remove the extra cells
bord=  Border(left=Side(border_style=None),right=Side(border_style=None),top=Side(border_style=None),bottom=Side(border_style=None))
for x in range(1,ws.max_row+1):
	ws['G' + str(x)].value=" "
	ws['H' + str(x)].value=" "
	ws['I' + str(x)].value=" "
	ws['J' + str(x)].value=" "
	ws['K' + str(x)].value=" "
	ws['L' + str(x)].value=" "
	ws['M' + str(x)].value=" "
	ws['N' + str(x)].value=" "
	ws['O' + str(x)].value=" "
	ws['G' + str(x)].border = bord
	ws['H' + str(x)].border = bord
	ws['I' + str(x)].border = bord
	ws['J' + str(x)].border = bord
	ws['K' + str(x)].border = bord
	ws['L' + str(x)].border = bord
	ws['M' + str(x)].border = bord
	ws['N' + str(x)].border = bord
	ws['O' + str(x)].border = bord


#save file finally
wb2.save("/var/www/html/CRT/HSBC/py/changed1.xlsx")


data_xls = pd.read_excel('/var/www/html/CRT/HSBC/py/changed1.xlsx', names[0], index_col=False, header=0)
data_xls.to_csv('/var/www/html/CRT/HSBC/py/your_csv.csv', header=None ,index=False,  date_format='%m/%d/%Y')

i=0
f = open('/var/www/html/CRT/HSBC/py/your_csv.csv', 'r')
linelist = f.readlines()
f.close

# Re-open file here
f2 = open('/var/www/html/CRT/HSBC/py/your_csv.csv', 'w')
for line in linelist:
		#if line.__contains__("abvp"):
		line=str(line.replace(" ",""))
		line = line.strip('\n')
		line = line.strip(',')
		print line
		f2.write(line+'\n')
		i=i+1
f2.close()


import os
os.remove('/var/www/html/CRT/HSBC/py/'+filename)
os.remove('/var/www/html/CRT/HSBC/py/changed1.xlsx')
