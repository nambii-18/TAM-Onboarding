
from openpyxl import Workbook
from openpyxl import load_workbook
import re
from openpyxl.styles import NamedStyle, Font, colors, Border, Side
from openpyxl.styles.numbers import NumberFormat
wb = Workbook()
import sys
import pandas as pd

from datetime import datetime

# HSBC  ticket-number    status    priority    date     engineer

filename = sys.argv[1]
#print filename + "sdssd"
#load sheets
wb2 = load_workbook('/var/www/html/CRT/HSBC/py/'+filename)
names = wb2.get_sheet_names()
ws = wb2.get_sheet_by_name(names[0])
#print ws.max_row

#set style
date_style = NamedStyle(name='datetime', number_format='MM/DD/YYYY')
#here you iterate over the rows in the specific column
#print([re.sub(u'(\u2018|\u2019)', "'", str(ws[x][5].value)) for x in range(2,ws.max_row)])
'''
B = ws['B']
print B[1].value
ft = Font(color=colors.RED)
B[1].font = ft
B[1].value = "Changed by me"
'''

#format columns - important
for x in range(8,ws.max_row+1):
    #print x,ws['C' + str(x)].value
    try:
        ws['C' + str(x)].value = ws['B' + str(x)].value
        ws['B' + str(x)].value =  ws['A' + str(x)].value
        ws['A' + str(x)].value =  ws['C' + str(x)].value
        ws['C' + str(x)].value =  ws['G' + str(x)].value
        #print ws['C' + str(x)].value
        if ws['D' + str(x)].value == "Minor":
            ws['D' + str(x)].value= 3
        elif ws['D' + str(x)].value == "Critical":
            ws['D' + str(x)].value= 1
        elif  ws['D' + str(x)].value == "Major":
            ws['D' + str(x)].value= 2
        elif  ws['D' + str(x)].value == "Notice":
            ws['D' + str(x)].value= 4


        ws['E' + str(x)].value = ws['F' + str(x)].value
        ws['E' + str(x)].style = date_style

        #date time format conversion
        #print ws['E' + str(x)].value
        oldformat = str(ws['E' + str(x)].value)
        oldformat = oldformat[0:10]
        #print oldformat
        oldformat= oldformat[5:10]+"-"+oldformat[0:4]
        ws['E' + str(x)].value=oldformat

        ws['F' + str(x)].value = ws['H' + str(x)].value


        ws['G' + str(x)].value=" "
        ws['H' + str(x)].value=" "
        ws['I' + str(x)].value=" "
        ws['J' + str(x)].value=" "
        ws['K' + str(x)].value=" "
        ws['L' + str(x)].value=" "
        ws['M' + str(x)].value=" "
        ws['N' + str(x)].value=" "
        ws['O' + str(x)].value=" "
    except Exception, e:
        print e
    finally:
        pass

#remove the extra cells
bord=  Border(left=Side(border_style=None),right=Side(border_style=None),top=Side(border_style=None),bottom=Side(border_style=None))
for x in range(8,ws.max_row+1):
    ws['G' + str(x)].value=" "
    ws['H' + str(x)].value=" "
    ws['I' + str(x)].value=" "
    ws['J' + str(x)].value=" "
    ws['K' + str(x)].value=" "
    ws['L' + str(x)].value=" "
    ws['M' + str(x)].value=" "
    ws['N' + str(x)].value=" "
    ws['O' + str(x)].value=" "
    ws['G' + str(x)].border = bord
    ws['H' + str(x)].border = bord
    ws['I' + str(x)].border = bord
    ws['J' + str(x)].border = bord
    ws['K' + str(x)].border = bord
    ws['L' + str(x)].border = bord
    ws['M' + str(x)].border = bord
    ws['N' + str(x)].border = bord
    ws['O' + str(x)].border = bord


#save file finally
wb2.save("/var/www/html/CRT/HSBC/py/changed.xlsx")


data_xls = pd.read_excel('/var/www/html/CRT/HSBC/py/changed.xlsx', names[0], index_col=0, header=0)
#data_xls[5] =  pd.to_datetime(data_xls[5], format='%m/%d/%Y')
data_xls.to_csv('/var/www/html/CRT/HSBC/py/cmsp_csv.csv',   date_format='%Y-%m-%d')

i=0
f = open('/var/www/html/CRT/HSBC/py/cmsp_csv.csv', 'r')
linelist = f.readlines()
f.close

# Re-open file here
f2 = open('/var/www/html/CRT/HSBC/py/cmsp_csv.csv', 'w')
for line in linelist:
        if line.__contains__("HSBC"):
            line=str(line.replace(" ",""))
            line = line.strip('\n')
            line = line.strip(',')
            print line
            f2.write(line+'\n')
            i=i+1
f2.close()


import os
os.remove('C:/wamp64/www/CRT/HSBC/py/'+filename)
os.remove('C:/wamp64/www/CRT/HSBC/py/changed.xlsx')
